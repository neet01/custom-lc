from __future__ import annotations

import ast
import base64
import csv
import io
import math
import re
from copy import copy
from datetime import datetime
from statistics import median
from typing import Any, Dict, List, Optional, Tuple

from fastapi import FastAPI, HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.formula.translate import Translator
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import range_boundaries
from pydantic import BaseModel, Field


SUPPORTED_OUTPUT_FORMATS = {"xlsx", "csv"}
SUPPORTED_OPERATION_TYPES = {
    "add_column",
    "add_row",
    "update_cells",
    "sort_rows",
    "add_totals_row",
    "reorder_rows",
    "merge_sheets",
    "split_sheet",
}
DEFAULT_REDACTION_TEXT = "[REDACTED]"
HEADER_PROBE_ROWS = 10
AGGREGATION_FUNCTIONS = {"sum", "average", "min", "max", "count", "counta"}


class UnsupportedOperationError(Exception):
    pass


class WorkerRequest(BaseModel):
    bufferBase64: str = Field(..., description="Base64 encoded workbook bytes")
    sourceFilename: str
    maxPreviewRows: int = 5
    removeColumns: List[str] = Field(default_factory=list)
    keepColumns: List[str] = Field(default_factory=list)
    redactColumns: List[str] = Field(default_factory=list)
    redactionText: str = DEFAULT_REDACTION_TEXT
    sheetNames: List[str] = Field(default_factory=list)
    outputFormat: Optional[str] = None
    operations: List[Dict[str, Any]] = Field(default_factory=list)


class HealthResponse(BaseModel):
    ok: bool
    engine: str


app = FastAPI(title="LibreChat Spreadsheet Worker", version="0.1.0")


def decode_workbook_buffer(payload: WorkerRequest) -> bytes:
    try:
        return base64.b64decode(payload.bufferBase64)
    except Exception as exc:  # pragma: no cover - defensive
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Invalid workbook payload",
                "code": "INVALID_BUFFER",
                "details": str(exc),
            },
        ) from exc


def normalize_scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def normalize_scalar_preserve_type(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def normalize_column_name(value: Any) -> str:
    return normalize_scalar(value).lower()


def coerce_number(value: Any) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_scalar(value)
    if text == "":
        return None
    normalized = text.replace(",", "").replace("$", "").replace("€", "").replace("£", "")
    if normalized.endswith("%"):
        normalized = normalized[:-1]
    try:
        return float(normalized)
    except ValueError:
        return None


def is_currency_format(value: Any, number_format: str) -> bool:
    if isinstance(value, (int, float)) and any(symbol in (number_format or "") for symbol in ("$", "€", "£")):
        return True
    return any(symbol in normalize_scalar(value) for symbol in ("$", "€", "£"))


def is_percentage_format(value: Any, number_format: str) -> bool:
    if isinstance(value, (int, float)) and "%" in (number_format or ""):
        return True
    return "%" in normalize_scalar(value)


def is_date_like(value: Any, number_format: str) -> bool:
    if isinstance(value, datetime):
        return True
    format_value = normalize_scalar(number_format).lower()
    return any(token in format_value for token in ("yy", "mm", "dd", "hh", "ss"))


def detect_header_row_index(worksheet) -> int:
    max_probe_row = min(max(worksheet.max_row, 1), HEADER_PROBE_ROWS)
    best_row = 1
    best_score = -1

    for row_index in range(1, max_probe_row + 1):
        score = 0
        for column_index in range(1, worksheet.max_column + 1):
            if normalize_scalar(worksheet.cell(row=row_index, column=column_index).value) != "":
                score += 1
        if score > best_score:
            best_score = score
            best_row = row_index

    return best_row


def get_headers(worksheet, header_row_index: int) -> List[str]:
    headers: List[str] = []
    for column_index in range(1, worksheet.max_column + 1):
        value = normalize_scalar(worksheet.cell(row=header_row_index, column=column_index).value)
        headers.append(value)
    while headers and headers[-1] == "":
        headers.pop()
    return headers


def get_sheet_preview(worksheet, header_row_index: int, headers: List[str], max_preview_rows: int):
    preview_rows: List[Dict[str, Any]] = []
    if not headers:
        return preview_rows

    max_row = min(worksheet.max_row, header_row_index + max_preview_rows)
    for row_index in range(header_row_index + 1, max_row + 1):
        row_payload: Dict[str, Any] = {}
        is_non_empty = False
        for column_index, header in enumerate(headers, start=1):
            cell_value = worksheet.cell(row=row_index, column=column_index).value
            normalized_value = normalize_scalar_preserve_type(cell_value)
            row_payload[header or f"Column {column_index}"] = normalized_value
            if normalize_scalar(normalized_value) != "":
                is_non_empty = True
        if is_non_empty:
            preview_rows.append(row_payload)

    return preview_rows


def get_formula_samples(worksheet, limit: int = 20):
    formulas = []
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append({"cell": cell.coordinate, "formula": cell.value})
                if len(formulas) >= limit:
                    return formulas
    return formulas


def detect_tables(worksheet, header_row_index: int, headers: List[str]):
    detected = []

    if getattr(worksheet, "tables", None):
        for table in worksheet.tables.values():
            detected.append(
                {
                    "name": table.name,
                    "ref": table.ref,
                    "source": "excel_table",
                }
            )

    if headers:
        detected.append(
            {
                "name": f"{worksheet.title}_main",
                "ref": f"A{header_row_index}:{column_letter(len(headers))}{worksheet.max_row}",
                "source": "header_heuristic",
            }
        )

    return detected


def build_column_profiles(worksheet, header_row_index: int, headers: List[str]):
    profiles = []

    for column_index, header in enumerate(headers, start=1):
        sample_values = []
        number_format_samples = []
        non_empty_count = 0
        numeric_count = 0
        date_count = 0
        formula_count = 0
        currency_like_count = 0
        percentage_like_count = 0
        numeric_values: List[float] = []

        for row_index in range(header_row_index + 1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            value = cell.value
            normalized_value = normalize_scalar(value)
            if normalized_value == "":
                continue

            non_empty_count += 1
            if len(sample_values) < 5:
                sample_values.append(normalize_scalar_preserve_type(value))

            if cell.number_format and len(number_format_samples) < 3:
                if cell.number_format not in number_format_samples:
                    number_format_samples.append(cell.number_format)

            if isinstance(value, str) and value.startswith("="):
                formula_count += 1

            if is_date_like(value, cell.number_format):
                date_count += 1

            numeric_value = coerce_number(value)
            if numeric_value is not None:
                numeric_count += 1
                numeric_values.append(numeric_value)
                if is_currency_format(value, cell.number_format):
                    currency_like_count += 1
                if is_percentage_format(value, cell.number_format):
                    percentage_like_count += 1

        inferred_type = "empty"
        if non_empty_count > 0:
            if date_count == non_empty_count:
                inferred_type = "date"
            elif numeric_count == non_empty_count and currency_like_count > 0:
                inferred_type = "currency"
            elif numeric_count == non_empty_count and percentage_like_count > 0:
                inferred_type = "percentage"
            elif numeric_count == non_empty_count:
                inferred_type = "number"
            elif formula_count == non_empty_count:
                inferred_type = "formula"
            elif numeric_count > 0 and numeric_count >= max(2, math.ceil(non_empty_count * 0.6)):
                inferred_type = "mostly_numeric"
            else:
                inferred_type = "text"

        numeric_summary = None
        if numeric_values:
            numeric_summary = {
                "sum": round(sum(numeric_values), 6),
                "min": round(min(numeric_values), 6),
                "max": round(max(numeric_values), 6),
                "average": round(sum(numeric_values) / len(numeric_values), 6),
            }

        profiles.append(
            {
                "columnName": header or f"Column {column_index}",
                "inferredType": inferred_type,
                "nonEmptyCount": non_empty_count,
                "numericCount": numeric_count,
                "dateCount": date_count,
                "formulaCount": formula_count,
                "sampleValues": sample_values,
                "numberFormatSamples": number_format_samples,
                "numericSummary": numeric_summary,
            }
        )

    return profiles


def column_letter(column_index: int) -> str:
    value = ""
    current = column_index
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        value = chr(65 + remainder) + value
    return value


def inspect_workbook(buffer: bytes, source_filename: str, max_preview_rows: int):
    workbook = load_input_workbook(buffer, source_filename)
    if not workbook.sheetnames:
        raise ValueError("Spreadsheet does not contain any sheets")

    preview_row_count = max(1, min(int(max_preview_rows or 5), 10))
    sheets = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        header_row_index = detect_header_row_index(worksheet)
        headers = get_headers(worksheet, header_row_index)
        preview_rows = get_sheet_preview(worksheet, header_row_index, headers, preview_row_count)
        formulas = get_formula_samples(worksheet)
        merged_ranges = [str(merge) for merge in worksheet.merged_cells.ranges]
        detected_tables = detect_tables(worksheet, header_row_index, headers)
        column_profiles = build_column_profiles(worksheet, header_row_index, headers)

        sheets.append(
            {
                "sheetName": sheet_name,
                "rowCount": max(worksheet.max_row - header_row_index, 0),
                "columnCount": len(headers),
                "columns": headers,
                "previewRows": preview_rows,
                "headerRowIndex": header_row_index,
                "worksheetDimensions": {
                    "maxRow": worksheet.max_row,
                    "maxColumn": worksheet.max_column,
                    "dimension": worksheet.calculate_dimension(),
                },
                "formulaCount": len(formulas),
                "formulaSamples": formulas,
                "mergedCellCount": len(merged_ranges),
                "mergedCells": merged_ranges[:50],
                "detectedTables": detected_tables,
                "columnProfiles": column_profiles,
            }
        )

    return {
        "engine": "python_worker",
        "filename": source_filename,
        "sheetCount": len(sheets),
        "sheets": sheets,
    }


def workbook_output_format(source_filename: str, requested: Optional[str]) -> str:
    if requested:
        next_format = requested.lower()
        if next_format not in SUPPORTED_OUTPUT_FORMATS:
            raise ValueError(f"Unsupported output format: {requested}")
        return next_format

    return "csv" if source_filename.lower().endswith(".csv") else "xlsx"


def build_output_filename(source_filename: str, output_format: str) -> str:
    base_name = re.sub(r"\.[^.]+$", "", source_filename or "spreadsheet")
    safe_base = base_name or "spreadsheet"
    return f"{safe_base}-transformed.{output_format}"


def normalize_string_list(values: List[Any]) -> List[str]:
    normalized = []
    seen = set()
    for value in values or []:
        text = normalize_scalar(value)
        if text and text not in seen:
            seen.add(text)
            normalized.append(text)
    return normalized


def normalize_operations(operations: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    normalized = []
    for operation in operations or []:
        if not isinstance(operation, dict):
            continue
        next_operation = {**operation, "type": normalize_scalar(operation.get("type"))}
        if next_operation["type"]:
            normalized.append(next_operation)
    return normalized


def load_input_workbook(buffer: bytes, source_filename: str):
    extension = source_filename.lower().rsplit(".", 1)[-1] if "." in source_filename else ""
    if extension == "csv":
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sanitize_sheet_name(
            re.sub(r"\.[^.]+$", "", source_filename or "Sheet1"),
            "Sheet1",
        )
        text_stream = io.StringIO(buffer.decode("utf-8-sig"))
        reader = csv.reader(text_stream)
        for row in reader:
            worksheet.append(row)
        return workbook

    return load_workbook(io.BytesIO(buffer), data_only=False, keep_vba=True)


def get_header_map(headers: List[str]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for index, header in enumerate(headers, start=1):
        normalized = normalize_column_name(header)
        if normalized and normalized not in mapping:
            mapping[normalized] = index
    return mapping


def resolve_target_sheets(workbook, sheet_names: List[str]) -> List[str]:
    if not sheet_names:
        return list(workbook.sheetnames)

    target = [name for name in sheet_names if name in workbook.sheetnames]
    if not target:
        raise ValueError("None of the requested sheet names were found in the spreadsheet")
    return target


def copy_cell_format(source_cell, target_cell):
    if source_cell is None:
        return
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.font:
        target_cell.font = copy(source_cell.font)
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.border:
        target_cell.border = copy(source_cell.border)
    if source_cell.protection:
        target_cell.protection = copy(source_cell.protection)


def copy_cell_contents(source_cell: Cell, target_cell: Cell, translate_formula: bool = False):
    if source_cell is None:
        target_cell.value = None
        return

    copy_cell_format(source_cell, target_cell)

    source_value = source_cell.value
    if (
        translate_formula
        and isinstance(source_value, str)
        and source_value.startswith("=")
        and source_cell.coordinate != target_cell.coordinate
    ):
        try:
            target_cell.value = Translator(source_value, origin=source_cell.coordinate).translate_formula(
                target_cell.coordinate
            )
        except Exception:
            target_cell.value = source_value
    else:
        target_cell.value = source_value

    if source_cell.hyperlink:
        target_cell._hyperlink = copy(source_cell.hyperlink)
    if source_cell.comment:
        target_cell.comment = copy(source_cell.comment)


def snapshot_cell(source_cell: Cell):
    return {
        "coordinate": source_cell.coordinate,
        "value": source_cell.value,
        "style": copy(source_cell._style) if source_cell.has_style else None,
        "number_format": source_cell.number_format,
        "font": copy(source_cell.font) if source_cell.font else None,
        "fill": copy(source_cell.fill) if source_cell.fill else None,
        "alignment": copy(source_cell.alignment) if source_cell.alignment else None,
        "border": copy(source_cell.border) if source_cell.border else None,
        "protection": copy(source_cell.protection) if source_cell.protection else None,
        "hyperlink": copy(source_cell.hyperlink) if source_cell.hyperlink else None,
        "comment": copy(source_cell.comment) if source_cell.comment else None,
    }


def copy_snapshot_to_cell(snapshot: Dict[str, Any], target_cell: Cell, translate_formula: bool = False):
    target_cell._style = copy(snapshot["style"]) if snapshot.get("style") else target_cell._style
    if snapshot.get("number_format"):
        target_cell.number_format = snapshot["number_format"]
    if snapshot.get("font"):
        target_cell.font = copy(snapshot["font"])
    if snapshot.get("fill"):
        target_cell.fill = copy(snapshot["fill"])
    if snapshot.get("alignment"):
        target_cell.alignment = copy(snapshot["alignment"])
    if snapshot.get("border"):
        target_cell.border = copy(snapshot["border"])
    if snapshot.get("protection"):
        target_cell.protection = copy(snapshot["protection"])

    source_value = snapshot.get("value")
    source_coordinate = snapshot.get("coordinate") or target_cell.coordinate
    if (
        translate_formula
        and isinstance(source_value, str)
        and source_value.startswith("=")
        and source_coordinate != target_cell.coordinate
    ):
        try:
            target_cell.value = Translator(source_value, origin=source_coordinate).translate_formula(
                target_cell.coordinate
            )
        except Exception:
            target_cell.value = source_value
    else:
        target_cell.value = source_value

    target_cell._hyperlink = copy(snapshot["hyperlink"]) if snapshot.get("hyperlink") else None
    target_cell.comment = copy(snapshot["comment"]) if snapshot.get("comment") else None


def sanitize_sheet_name(value: Any, fallback: str = "Sheet") -> str:
    normalized = re.sub(r'[\\/?*\[\]:]', " ", str(value or "")).strip() or fallback
    return normalized[:31]


def ensure_unique_sheet_name(existing_sheet_names: List[str], desired_name: str) -> str:
    used_names = set(existing_sheet_names)
    if desired_name not in used_names:
        return desired_name

    suffix = 2
    while suffix < 1000:
        candidate = sanitize_sheet_name(f"{desired_name[:27]} {suffix}", desired_name)
        if candidate not in used_names:
            return candidate
        suffix += 1

    raise ValueError(f'Could not generate a unique sheet name from "{desired_name}"')


def copy_sheet_properties(source_worksheet, target_worksheet):
    target_worksheet.sheet_format = copy(source_worksheet.sheet_format)
    target_worksheet.sheet_properties = copy(source_worksheet.sheet_properties)
    target_worksheet.page_margins = copy(source_worksheet.page_margins)
    target_worksheet.page_setup = copy(source_worksheet.page_setup)
    target_worksheet.print_options = copy(source_worksheet.print_options)
    target_worksheet.sheet_view = copy(source_worksheet.sheet_view)
    target_worksheet.freeze_panes = source_worksheet.freeze_panes

    for column_key, column_dimension in source_worksheet.column_dimensions.items():
        target_dimension = target_worksheet.column_dimensions[column_key]
        for attribute in ("width", "hidden", "bestFit", "outlineLevel", "collapsed", "style", "min", "max"):
            value = getattr(column_dimension, attribute, None)
            if value is not None:
                setattr(target_dimension, attribute, value)


def copy_row_dimension(source_worksheet, source_row_index: int, target_worksheet, target_row_index: int):
    source_dimension = source_worksheet.row_dimensions[source_row_index]
    target_dimension = target_worksheet.row_dimensions[target_row_index]
    for attribute in ("height", "hidden", "outlineLevel", "collapsed", "style", "ht", "customFormat", "customHeight"):
        value = getattr(source_dimension, attribute, None)
        if value is not None:
            setattr(target_dimension, attribute, value)


def clear_worksheet(worksheet):
    if worksheet.max_row:
        worksheet.delete_rows(1, worksheet.max_row)
    if worksheet.max_column:
        worksheet.delete_cols(1, worksheet.max_column)
    worksheet.freeze_panes = None
    worksheet.auto_filter.ref = None
    worksheet.merged_cells.ranges = set()


def build_autofilter_ref(column_count: int, row_count: int) -> Optional[str]:
    if column_count <= 0 or row_count <= 1:
        return None
    return f"A1:{column_letter(column_count)}{row_count}"


def copy_selected_rows_to_sheet(
    source_worksheet,
    target_worksheet,
    source_row_indexes: List[int],
    translate_formula_data_rows: Optional[set[int]] = None,
):
    row_mapping: Dict[int, int] = {}
    for target_row_index, source_row_index in enumerate(source_row_indexes, start=1):
        row_mapping[source_row_index] = target_row_index
        copy_row_dimension(source_worksheet, source_row_index, target_worksheet, target_row_index)
        for column_index in range(1, source_worksheet.max_column + 1):
            snapshot = snapshot_cell(source_worksheet.cell(row=source_row_index, column=column_index))
            copy_snapshot_to_cell(
                snapshot,
                target_worksheet.cell(row=target_row_index, column=column_index),
                translate_formula=(
                    translate_formula_data_rows is not None and source_row_index in translate_formula_data_rows
                ),
            )

    for merged_range in source_worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        selected_rows = set(range(min_row, max_row + 1))
        if not selected_rows.issubset(row_mapping.keys()):
            continue
        target_worksheet.merge_cells(
            start_row=row_mapping[min_row],
            start_column=min_col,
            end_row=row_mapping[max_row],
            end_column=max_col,
        )

    return row_mapping


def delete_columns_by_index(worksheet, column_indexes: List[int]):
    for column_index in sorted(column_indexes, reverse=True):
        worksheet.delete_cols(column_index, 1)


def apply_legacy_column_transforms(
    worksheet,
    keep_columns: List[str],
    remove_columns: List[str],
    redact_columns: List[str],
    redaction_text: str,
):
    header_row_index = detect_header_row_index(worksheet)
    headers = get_headers(worksheet, header_row_index)
    header_map = get_header_map(headers)
    matched = {"keep": [], "remove": [], "redact": []}

    included_headers = list(headers)
    if keep_columns:
        keep_set = {normalize_column_name(value) for value in keep_columns}
        included_headers = [header for header in headers if normalize_column_name(header) in keep_set]
        matched["keep"] = [header for header in headers if normalize_column_name(header) in keep_set]

    remove_set = {normalize_column_name(value) for value in remove_columns}
    matched["remove"] = [header for header in included_headers if normalize_column_name(header) in remove_set]
    final_headers = [header for header in included_headers if normalize_column_name(header) not in remove_set]

    if keep_columns or remove_columns:
        delete_candidates = []
        for index, header in enumerate(headers, start=1):
            normalized = normalize_column_name(header)
            if keep_columns and normalized not in {normalize_column_name(value) for value in keep_columns}:
                delete_candidates.append(index)
            elif normalized in remove_set:
                delete_candidates.append(index)
        if len(delete_candidates) == len(headers) and headers:
            raise ValueError(f'All columns were removed from sheet "{worksheet.title}"')
        delete_columns_by_index(worksheet, delete_candidates)
        headers = get_headers(worksheet, header_row_index)
        header_map = get_header_map(headers)

    redact_set = {normalize_column_name(value) for value in redact_columns}
    matched["redact"] = [header for header in headers if normalize_column_name(header) in redact_set]
    if redact_set:
        for column_name in matched["redact"]:
            column_index = header_map[normalize_column_name(column_name)]
            for row_index in range(header_row_index + 1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_index, column=column_index)
                if normalize_scalar(cell.value) == "":
                    continue
                cell.value = redaction_text
                if cell.fill == PatternFill():
                    cell.fill = PatternFill(fill_type="solid", start_color="FFF4CCCC", end_color="FFF4CCCC")

    return {
        "sheetName": worksheet.title,
        "headerRowIndex": header_row_index,
        "originalColumnCount": len(final_headers) + len(matched["remove"]) if (keep_columns or remove_columns) else len(headers),
        "outputColumnCount": len(headers),
        "rowCount": max(worksheet.max_row - header_row_index, 0),
        "keptColumns": matched["keep"] or headers,
        "removedColumns": matched["remove"],
        "redactedColumns": matched["redact"],
    }


def build_row_object(worksheet, headers: List[str], row_index: int) -> Dict[str, Any]:
    row_object: Dict[str, Any] = {"__row_number": row_index}
    for column_index, header in enumerate(headers, start=1):
        if not header:
            continue
        row_object[header] = worksheet.cell(row=row_index, column=column_index).value
    return row_object


def formula_cell_reference(header_map: Dict[str, int], column_name: str, row_number: int) -> str:
    column_index = header_map.get(normalize_column_name(column_name))
    if column_index is None:
        raise ValueError(f'Formula references unknown column "{column_name}"')
    return f"{column_letter(column_index)}{row_number}"


def build_formula_from_template(template: str, headers: List[str], row_number: int) -> str:
    header_map = get_header_map(headers)
    formula_body = normalize_scalar(template)
    if not formula_body:
        raise ValueError("Formula template is empty")
    if formula_body.startswith("="):
        formula_body = formula_body[1:]

    def replace(match: re.Match[str]) -> str:
        return formula_cell_reference(header_map, match.group(1), row_number)

    return "=" + re.sub(r"{{\s*([^}]+?)\s*}}", replace, formula_body)


def safe_roundup(value, precision=0):
    factor = 10 ** int(precision or 0)
    return math.ceil(float(value) * factor) / factor


def safe_rounddown(value, precision=0):
    factor = 10 ** int(precision or 0)
    return math.floor(float(value) * factor) / factor


SAFE_FUNCTIONS = {
    "SUM": lambda *values: sum(float(value or 0) for value in values),
    "AVERAGE": lambda *values: (sum(float(value or 0) for value in values) / len(values)) if values else 0,
    "MIN": lambda *values: min(values) if values else 0,
    "MAX": lambda *values: max(values) if values else 0,
    "ABS": abs,
    "ROUND": round,
    "IF": lambda condition, when_true, when_false: when_true if condition else when_false,
    "AND": lambda *values: all(values),
    "OR": lambda *values: any(values),
    "NOT": lambda value: not value,
    "CONCAT": lambda *values: "".join("" if value is None else str(value) for value in values),
    "CONCATENATE": lambda *values: "".join("" if value is None else str(value) for value in values),
    "COUNT": lambda *values: sum(1 for value in values if isinstance(value, (int, float)) and not isinstance(value, bool)),
    "COUNTA": lambda *values: sum(1 for value in values if normalize_scalar(value) != ""),
    "MEDIAN": lambda *values: median([float(value) for value in values]) if values else 0,
    "LEN": lambda value: len(str(value or "")),
    "LOWER": lambda value: str(value or "").lower(),
    "UPPER": lambda value: str(value or "").upper(),
    "TRIM": lambda value: str(value or "").strip(),
    "ROUNDUP": safe_roundup,
    "ROUNDDOWN": safe_rounddown,
}


class SafeExpressionEvaluator(ast.NodeVisitor):
    ALLOWED_NODES = (
        ast.Expression,
        ast.BinOp,
        ast.UnaryOp,
        ast.BoolOp,
        ast.Compare,
        ast.Call,
        ast.Name,
        ast.Load,
        ast.Constant,
        ast.Add,
        ast.Sub,
        ast.Mult,
        ast.Div,
        ast.Mod,
        ast.Pow,
        ast.USub,
        ast.UAdd,
        ast.Eq,
        ast.NotEq,
        ast.Lt,
        ast.LtE,
        ast.Gt,
        ast.GtE,
        ast.And,
        ast.Or,
        ast.Not,
        ast.IfExp,
    )

    def __init__(self, context: Dict[str, Any]):
        self.context = context

    def visit(self, node):  # type: ignore[override]
        if not isinstance(node, self.ALLOWED_NODES):
            raise ValueError(f"Unsupported expression element: {node.__class__.__name__}")
        return super().visit(node)

    def visit_Expression(self, node: ast.Expression):
        return self.visit(node.body)

    def visit_Constant(self, node: ast.Constant):
        return node.value

    def visit_Name(self, node: ast.Name):
        if node.id in self.context:
            return self.context[node.id]
        raise ValueError(f'Unknown expression variable "{node.id}"')

    def visit_BinOp(self, node: ast.BinOp):
        left = self.visit(node.left)
        right = self.visit(node.right)
        if isinstance(node.op, ast.Add):
            return left + right
        if isinstance(node.op, ast.Sub):
            return left - right
        if isinstance(node.op, ast.Mult):
            return left * right
        if isinstance(node.op, ast.Div):
            return left / right
        if isinstance(node.op, ast.Mod):
            return left % right
        if isinstance(node.op, ast.Pow):
            return left ** right
        raise ValueError("Unsupported binary operation")

    def visit_UnaryOp(self, node: ast.UnaryOp):
        operand = self.visit(node.operand)
        if isinstance(node.op, ast.USub):
            return -operand
        if isinstance(node.op, ast.UAdd):
            return +operand
        if isinstance(node.op, ast.Not):
            return not operand
        raise ValueError("Unsupported unary operation")

    def visit_BoolOp(self, node: ast.BoolOp):
        values = [self.visit(value) for value in node.values]
        if isinstance(node.op, ast.And):
            return all(values)
        if isinstance(node.op, ast.Or):
            return any(values)
        raise ValueError("Unsupported boolean operation")

    def visit_Compare(self, node: ast.Compare):
        left = self.visit(node.left)
        for operator, comparator in zip(node.ops, node.comparators):
            right = self.visit(comparator)
            if isinstance(operator, ast.Eq) and not (left == right):
                return False
            if isinstance(operator, ast.NotEq) and not (left != right):
                return False
            if isinstance(operator, ast.Lt) and not (left < right):
                return False
            if isinstance(operator, ast.LtE) and not (left <= right):
                return False
            if isinstance(operator, ast.Gt) and not (left > right):
                return False
            if isinstance(operator, ast.GtE) and not (left >= right):
                return False
            left = right
        return True

    def visit_Call(self, node: ast.Call):
        if not isinstance(node.func, ast.Name):
            raise ValueError("Only named functions are allowed in expressions")
        function_name = node.func.id.upper()
        function = SAFE_FUNCTIONS.get(function_name)
        if function is None:
            raise ValueError(f'Unsupported function "{function_name}"')
        args = [self.visit(arg) for arg in node.args]
        return function(*args)

    def visit_IfExp(self, node: ast.IfExp):
        return self.visit(node.body) if self.visit(node.test) else self.visit(node.orelse)


def sanitize_expression(expression: str) -> str:
    next_expression = normalize_scalar(expression)
    if next_expression.startswith("="):
        next_expression = next_expression[1:].strip()
    next_expression = next_expression.replace("“", '"').replace("”", '"').replace("‘", "'").replace("’", "'")
    next_expression = next_expression.replace(";", ",")
    next_expression = re.sub(r"\bTRUE\b", "True", next_expression, flags=re.IGNORECASE)
    next_expression = re.sub(r"\bFALSE\b", "False", next_expression, flags=re.IGNORECASE)
    next_expression = next_expression.replace("<>", "!=")
    next_expression = re.sub(r"(?<![<>!=])=(?!=)", "==", next_expression)
    next_expression = re.sub(r"(\d+(?:\.\d+)?)%", r"(\1 / 100)", next_expression)
    return next_expression


def evaluate_expression_template(template: str, row_object: Dict[str, Any]) -> Any:
    variable_index = 0
    context: Dict[str, Any] = {"row_number": row_object.get("__row_number", 0)}

    def replace(match: re.Match[str]) -> str:
        nonlocal variable_index
        column_name = normalize_scalar(match.group(1))
        token = f"col_{variable_index}"
        variable_index += 1
        context[token] = row_object.get(column_name, "")
        return token

    prepared = re.sub(r"{{\s*([^}]+?)\s*}}", replace, template)
    parsed = ast.parse(sanitize_expression(prepared), mode="eval")
    return SafeExpressionEvaluator(context).visit(parsed)


def row_target_indexes(worksheet, headers: List[str], operation: Dict[str, Any], header_row_index: int) -> List[int]:
    if isinstance(operation.get("rowNumber"), int):
        target_row = header_row_index + operation["rowNumber"]
        if target_row > worksheet.max_row:
            raise ValueError(f'Row number {operation["rowNumber"]} is out of range for sheet "{worksheet.title}"')
        return [target_row]

    row_match = operation.get("rowMatch")
    if isinstance(row_match, dict) and row_match:
        header_map = get_header_map(headers)
        matching_rows: List[int] = []
        for row_index in range(header_row_index + 1, worksheet.max_row + 1):
            matches = True
            for column_name, expected in row_match.items():
                column_index = header_map.get(normalize_column_name(column_name))
                if column_index is None:
                    matches = False
                    break
                actual = worksheet.cell(row=row_index, column=column_index).value
                if normalize_scalar(actual) != normalize_scalar(expected):
                    matches = False
                    break
            if matches:
                matching_rows.append(row_index)
        return matching_rows

    return list(range(header_row_index + 1, worksheet.max_row + 1))


def column_insert_index(headers: List[str], operation: Dict[str, Any]) -> int:
    header_map = get_header_map(headers)
    before_column = normalize_scalar(operation.get("beforeColumn"))
    after_column = normalize_scalar(operation.get("afterColumn"))
    if before_column:
        column_index = header_map.get(normalize_column_name(before_column))
        if column_index is None:
            raise ValueError(f'Could not find beforeColumn "{before_column}"')
        return column_index
    if after_column:
        column_index = header_map.get(normalize_column_name(after_column))
        if column_index is None:
            raise ValueError(f'Could not find afterColumn "{after_column}"')
        return column_index + 1
    if isinstance(operation.get("index"), int):
        return max(1, min(len(headers) + 1, operation["index"] + 1))
    if normalize_scalar(operation.get("position")) == "start":
        return 1
    return len(headers) + 1


def apply_add_column(worksheet, operation: Dict[str, Any]):
    header_row_index = detect_header_row_index(worksheet)
    headers = get_headers(worksheet, header_row_index)
    column_name = normalize_scalar(operation.get("columnName"))
    if not column_name:
        raise ValueError("add_column operations require columnName")
    if normalize_column_name(column_name) in get_header_map(headers):
        raise ValueError(f'Sheet "{worksheet.title}" already contains column "{column_name}"')

    insert_index = column_insert_index(headers, operation)
    worksheet.insert_cols(insert_index, 1)

    header_cell = worksheet.cell(row=header_row_index, column=insert_index)
    template_cell = (
        worksheet.cell(row=header_row_index, column=max(1, insert_index - 1))
        if insert_index > 1
        else worksheet.cell(row=header_row_index, column=min(worksheet.max_column, insert_index + 1))
    )
    copy_cell_format(template_cell, header_cell)
    header_cell.value = column_name

    headers = get_headers(worksheet, header_row_index)
    for row_index in range(header_row_index + 1, worksheet.max_row + 1):
        row_object = build_row_object(worksheet, headers, row_index)
        target_cell = worksheet.cell(row=row_index, column=insert_index)
        template_row_cell = (
            worksheet.cell(row=row_index, column=max(1, insert_index - 1))
            if insert_index > 1
            else worksheet.cell(row=row_index, column=min(worksheet.max_column, insert_index + 1))
        )
        copy_cell_format(template_row_cell, target_cell)
        if normalize_scalar(operation.get("formula")):
            target_cell.value = build_formula_from_template(operation["formula"], headers, row_index)
        elif normalize_scalar(operation.get("expression")):
            target_cell.value = evaluate_expression_template(operation["expression"], row_object)
        elif "defaultValue" in operation:
            target_cell.value = operation.get("defaultValue")
        else:
            target_cell.value = ""


def apply_update_cells(worksheet, operation: Dict[str, Any]):
    header_row_index = detect_header_row_index(worksheet)
    headers = get_headers(worksheet, header_row_index)
    column_name = normalize_scalar(operation.get("columnName"))
    if not column_name:
        raise ValueError("update_cells operations require columnName")

    header_map = get_header_map(headers)
    column_index = header_map.get(normalize_column_name(column_name))
    if column_index is None:
        raise ValueError(f'Could not find column "{column_name}"')

    row_indexes = row_target_indexes(worksheet, headers, operation, header_row_index)
    if not row_indexes:
        raise ValueError(f'No rows matched update_cells operation on sheet "{worksheet.title}"')

    for row_index in row_indexes:
        row_object = build_row_object(worksheet, headers, row_index)
        cell = worksheet.cell(row=row_index, column=column_index)
        if normalize_scalar(operation.get("formula")):
            cell.value = build_formula_from_template(operation["formula"], headers, row_index)
        elif normalize_scalar(operation.get("expression")):
            cell.value = evaluate_expression_template(operation["expression"], row_object)
        elif "value" in operation:
            cell.value = operation.get("value")
        else:
            raise ValueError("update_cells operations require value, expression, or formula")


def apply_add_row(worksheet, operation: Dict[str, Any]):
    header_row_index = detect_header_row_index(worksheet)
    headers = get_headers(worksheet, header_row_index)
    values = operation.get("values")
    if not isinstance(values, dict):
        raise ValueError("add_row operations require a values object")

    insert_row = worksheet.max_row + 1
    if isinstance(operation.get("index"), int):
        insert_row = max(header_row_index + 1, min(worksheet.max_row + 1, header_row_index + operation["index"]))
        worksheet.insert_rows(insert_row, 1)
    elif normalize_scalar(operation.get("position")) == "start":
        insert_row = header_row_index + 1
        worksheet.insert_rows(insert_row, 1)

    header_map = get_header_map(headers)
    template_row = min(max(insert_row - 1, header_row_index), worksheet.max_row)
    inherit_formulas = operation.get("inheritFormulas", True) is not False
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=insert_row, column=column_index)
        template_cell = worksheet.cell(row=template_row, column=column_index) if template_row >= 1 else None
        if template_row >= 1 and template_row != insert_row:
            copy_cell_format(template_cell, cell)
        value = ""
        has_explicit_value = False
        for provided_header, provided_value in values.items():
            if normalize_column_name(provided_header) == normalize_column_name(header):
                value = provided_value
                has_explicit_value = True
                break
        if has_explicit_value:
            cell.value = value
        elif (
            inherit_formulas
            and template_cell is not None
            and isinstance(template_cell.value, str)
            and template_cell.value.startswith("=")
        ):
            copy_cell_contents(template_cell, cell, translate_formula=True)
        else:
            cell.value = value

    for provided_header in values.keys():
        if normalize_column_name(provided_header) not in header_map:
            raise ValueError(f'Could not find row value column "{provided_header}"')


def build_sort_spec(operation: Dict[str, Any]):
    if isinstance(operation.get("columns"), list) and operation["columns"]:
        specs = []
        for item in operation["columns"]:
            if not isinstance(item, dict):
                continue
            column_name = normalize_scalar(item.get("columnName"))
            if not column_name:
                continue
            specs.append(
                {
                    "columnName": column_name,
                    "direction": normalize_scalar(item.get("direction")) or "asc",
                    "numeric": bool(item.get("numeric")),
                }
            )
        if specs:
            return specs

    column_name = normalize_scalar(operation.get("columnName"))
    if not column_name:
        raise ValueError("sort_rows operations require columnName or columns")

    return [
        {
            "columnName": column_name,
            "direction": normalize_scalar(operation.get("direction")) or "asc",
            "numeric": bool(operation.get("numeric")),
        }
    ]


def build_sort_key(value: Any, numeric: bool):
    if numeric:
        number = coerce_number(value)
        return (number is None, number if number is not None else 0)

    text = normalize_scalar(value)
    return (text == "", text.lower())


def apply_sort_rows(worksheet, operation: Dict[str, Any]):
    header_row_index = detect_header_row_index(worksheet)
    headers = get_headers(worksheet, header_row_index)
    header_map = get_header_map(headers)
    sort_spec = build_sort_spec(operation)

    data_row_indexes = list(range(header_row_index + 1, worksheet.max_row + 1))
    row_snapshots = []
    for row_index in data_row_indexes:
        row_object = build_row_object(worksheet, headers, row_index)
        row_snapshots.append(
            {
                "rowIndex": row_index,
                "rowObject": row_object,
                "cells": [
                    snapshot_cell(worksheet.cell(row=row_index, column=column_index))
                    for column_index in range(1, worksheet.max_column + 1)
                ],
            }
        )

    for spec in reversed(sort_spec):
        column_name = spec["columnName"]
        if normalize_column_name(column_name) not in header_map:
            raise ValueError(f'Could not find sort column "{column_name}"')
        row_snapshots.sort(
            key=lambda snapshot: build_sort_key(snapshot["rowObject"].get(column_name), spec["numeric"]),
            reverse=spec["direction"].lower() == "desc",
        )

    for destination_row_index, snapshot in zip(data_row_indexes, row_snapshots):
        for column_index, source_cell in enumerate(snapshot["cells"], start=1):
            target_cell = worksheet.cell(row=destination_row_index, column=column_index)
            copy_snapshot_to_cell(source_cell, target_cell, translate_formula=True)


def apply_reorder_rows(worksheet, operation: Dict[str, Any]):
    ordered_row_numbers = operation.get("orderedRowNumbers")
    if not isinstance(ordered_row_numbers, list) or len(ordered_row_numbers) == 0:
        raise ValueError("reorder_rows operations require orderedRowNumbers")

    requested_indexes: List[int] = []
    seen_indexes = set()
    for value in ordered_row_numbers:
        if isinstance(value, bool):
            continue
        numeric_value = int(value) if isinstance(value, int) or (isinstance(value, float) and value.is_integer()) else None
        if numeric_value is None or numeric_value < 1 or numeric_value in seen_indexes:
            continue
        seen_indexes.add(numeric_value)
        requested_indexes.append(numeric_value)

    if len(requested_indexes) == 0:
        raise ValueError("reorder_rows orderedRowNumbers must contain positive integers")

    header_row_index = detect_header_row_index(worksheet)
    data_row_indexes = list(range(header_row_index + 1, worksheet.max_row + 1))
    row_snapshots = []
    for row_index in data_row_indexes:
        row_snapshots.append(
            {
                "rowIndex": row_index,
                "cells": [
                    snapshot_cell(worksheet.cell(row=row_index, column=column_index))
                    for column_index in range(1, worksheet.max_column + 1)
                ],
            }
        )

    for row_number in requested_indexes:
        if row_number > len(row_snapshots):
            raise ValueError(f'Row number {row_number} is out of range for sheet "{worksheet.title}"')

    requested_zero_indexes = [row_number - 1 for row_number in requested_indexes]
    reordered_snapshots = [row_snapshots[index] for index in requested_zero_indexes]

    if operation.get("appendRemaining", True) is not False:
        requested_set = set(requested_zero_indexes)
        reordered_snapshots.extend(
            snapshot
            for data_row_index, snapshot in enumerate(row_snapshots)
            if data_row_index not in requested_set
        )

    for destination_row_index, snapshot in zip(data_row_indexes, reordered_snapshots):
        for column_index, source_cell in enumerate(snapshot["cells"], start=1):
            target_cell = worksheet.cell(row=destination_row_index, column=column_index)
            copy_snapshot_to_cell(source_cell, target_cell, translate_formula=True)


def build_aggregation_formula(function_name: str, column_letter_value: str, start_row: int, end_row: int) -> str:
    if function_name == "sum":
        return f"=SUM({column_letter_value}{start_row}:{column_letter_value}{end_row})"
    if function_name == "average":
        return f"=AVERAGE({column_letter_value}{start_row}:{column_letter_value}{end_row})"
    if function_name == "min":
        return f"=MIN({column_letter_value}{start_row}:{column_letter_value}{end_row})"
    if function_name == "max":
        return f"=MAX({column_letter_value}{start_row}:{column_letter_value}{end_row})"
    if function_name == "count":
        return f"=COUNT({column_letter_value}{start_row}:{column_letter_value}{end_row})"
    if function_name == "counta":
        return f"=COUNTA({column_letter_value}{start_row}:{column_letter_value}{end_row})"
    raise ValueError(f'Unsupported aggregation function "{function_name}"')


def infer_numeric_total_columns(worksheet, headers: List[str], header_row_index: int):
    profiles = build_column_profiles(worksheet, header_row_index, headers)
    return [
        profile["columnName"]
        for profile in profiles
        if profile["inferredType"] in {"currency", "percentage", "number", "mostly_numeric"}
    ]


def apply_add_totals_row(worksheet, operation: Dict[str, Any]):
    header_row_index = detect_header_row_index(worksheet)
    headers = get_headers(worksheet, header_row_index)
    header_map = get_header_map(headers)
    if not headers:
        raise ValueError(f'Sheet "{worksheet.title}" does not have a usable header row')

    data_start_row = header_row_index + 1
    data_end_row = worksheet.max_row
    if data_end_row < data_start_row:
        raise ValueError(f'Sheet "{worksheet.title}" does not contain any data rows to total')

    insert_row = worksheet.max_row + 1
    worksheet.insert_rows(insert_row, 1)

    label_column_name = normalize_scalar(operation.get("labelColumn")) or headers[0]
    label_column_index = header_map.get(normalize_column_name(label_column_name))
    if label_column_index is None:
        raise ValueError(f'Could not find totals label column "{label_column_name}"')

    label = normalize_scalar(operation.get("label")) or "Total"
    summary_columns = operation.get("columns")
    if isinstance(summary_columns, list) and summary_columns:
        targets = []
        for item in summary_columns:
            if not isinstance(item, dict):
                continue
            column_name = normalize_scalar(item.get("columnName"))
            if not column_name:
                continue
            aggregation = normalize_scalar(item.get("function")) or "sum"
            if aggregation not in AGGREGATION_FUNCTIONS:
                raise ValueError(f'Unsupported totals aggregation "{aggregation}"')
            targets.append({"columnName": column_name, "function": aggregation})
    else:
        targets = [
            {"columnName": column_name, "function": "sum"}
            for column_name in infer_numeric_total_columns(worksheet, headers, header_row_index)
            if normalize_column_name(column_name) != normalize_column_name(label_column_name)
        ]

    if not targets:
        raise ValueError(f'Sheet "{worksheet.title}" does not have any numeric columns for a totals row')

    template_row = max(insert_row - 1, header_row_index)
    for column_index, header in enumerate(headers, start=1):
        target_cell = worksheet.cell(row=insert_row, column=column_index)
        template_cell = worksheet.cell(row=template_row, column=column_index)
        copy_cell_format(template_cell, target_cell)
        if column_index == label_column_index:
            target_cell.value = label
            if target_cell.font:
                target_cell.font = copy(target_cell.font)
                target_cell.font = Font(
                    name=target_cell.font.name,
                    size=target_cell.font.size,
                    bold=True,
                    italic=target_cell.font.italic,
                    vertAlign=target_cell.font.vertAlign,
                    underline=target_cell.font.underline,
                    strike=target_cell.font.strike,
                    color=target_cell.font.color,
                )
            continue
        target_cell.value = ""

    for target in targets:
        column_name = target["columnName"]
        column_index = header_map.get(normalize_column_name(column_name))
        if column_index is None:
            raise ValueError(f'Could not find totals column "{column_name}"')
        target_cell = worksheet.cell(row=insert_row, column=column_index)
        target_cell.value = build_aggregation_formula(
            target["function"],
            column_letter(column_index),
            data_start_row,
            insert_row - 1,
        )


def build_merge_headers(source_worksheets: List[Any], include_source_sheet_column: bool) -> List[str]:
    merged_headers: List[str] = []
    seen_headers = set()

    if include_source_sheet_column:
        merged_headers.append("Source Sheet")
        seen_headers.add(normalize_column_name("Source Sheet"))

    for worksheet in source_worksheets:
        header_row_index = detect_header_row_index(worksheet)
        headers = get_headers(worksheet, header_row_index)
        for header in headers:
            normalized = normalize_column_name(header)
            if not header or normalized in seen_headers:
                continue
            seen_headers.add(normalized)
            merged_headers.append(header)

    return merged_headers


def select_output_merge_sheet(workbook, source_sheet_names: List[str], desired_name: str):
    if desired_name in workbook.sheetnames:
        if desired_name not in source_sheet_names:
            raise ValueError(f'merge_sheets output sheet "{desired_name}" already exists')
        return workbook[desired_name], desired_name

    output_sheet_name = ensure_unique_sheet_name(list(workbook.sheetnames), desired_name)
    return workbook.create_sheet(title=output_sheet_name), output_sheet_name


def apply_merge_sheets(workbook, operation: Dict[str, Any]):
    source_sheet_names = normalize_string_list(operation.get("sourceSheets", []))
    if len(source_sheet_names) < 2:
        raise ValueError("merge_sheets operations require at least two sourceSheets")

    source_worksheets = []
    for sheet_name in source_sheet_names:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f'merge_sheets could not find source sheet "{sheet_name}"')
        source_worksheets.append(workbook[sheet_name])

    desired_name = sanitize_sheet_name(
        operation.get("outputSheetName") or f"{source_sheet_names[0]} Merged",
        "Merged Sheet",
    )
    output_worksheet, output_sheet_name = select_output_merge_sheet(
        workbook,
        source_sheet_names,
        desired_name,
    )

    include_source_sheet_column = operation.get("includeSourceSheetColumn", True) is not False
    merged_headers = build_merge_headers(source_worksheets, include_source_sheet_column)
    if not merged_headers:
        raise ValueError("merge_sheets could not determine any headers to merge")

    clear_worksheet(output_worksheet)
    copy_sheet_properties(source_worksheets[0], output_worksheet)

    header_templates: Dict[str, Dict[str, Any]] = {}
    for worksheet in source_worksheets:
        header_row_index = detect_header_row_index(worksheet)
        headers = get_headers(worksheet, header_row_index)
        for column_index, header in enumerate(headers, start=1):
            normalized = normalize_column_name(header)
            if header and normalized not in header_templates:
                header_templates[normalized] = snapshot_cell(
                    worksheet.cell(row=header_row_index, column=column_index)
                )

    if include_source_sheet_column and source_worksheets:
        first_header_row_index = detect_header_row_index(source_worksheets[0])
        header_templates[normalize_column_name("Source Sheet")] = snapshot_cell(
            source_worksheets[0].cell(row=first_header_row_index, column=1)
        )

    for column_index, header in enumerate(merged_headers, start=1):
        snapshot = header_templates.get(normalize_column_name(header))
        target_cell = output_worksheet.cell(row=1, column=column_index)
        if snapshot:
            copy_snapshot_to_cell(snapshot, target_cell, translate_formula=False)
        target_cell.value = header

    output_row_index = 2
    for worksheet in source_worksheets:
        header_row_index = detect_header_row_index(worksheet)
        headers = get_headers(worksheet, header_row_index)
        header_map = get_header_map(headers)
        for source_row_index in range(header_row_index + 1, worksheet.max_row + 1):
            row_has_content = False
            for source_column_index in range(1, worksheet.max_column + 1):
                if normalize_scalar(worksheet.cell(row=source_row_index, column=source_column_index).value) != "":
                    row_has_content = True
                    break
            if not row_has_content:
                continue

            copy_row_dimension(worksheet, source_row_index, output_worksheet, output_row_index)
            for output_column_index, header in enumerate(merged_headers, start=1):
                target_cell = output_worksheet.cell(row=output_row_index, column=output_column_index)
                if include_source_sheet_column and header == "Source Sheet":
                    template_snapshot = header_templates.get(normalize_column_name("Source Sheet"))
                    if template_snapshot:
                        copy_snapshot_to_cell(template_snapshot, target_cell, translate_formula=False)
                    target_cell.value = worksheet.title
                    continue

                source_column_index = header_map.get(normalize_column_name(header))
                if source_column_index is None:
                    target_cell.value = ""
                    continue

                snapshot = snapshot_cell(worksheet.cell(row=source_row_index, column=source_column_index))
                copy_snapshot_to_cell(snapshot, target_cell, translate_formula=False)
            output_row_index += 1

    auto_filter_ref = build_autofilter_ref(len(merged_headers), output_row_index - 1)
    if auto_filter_ref:
        output_worksheet.auto_filter.ref = auto_filter_ref

    if operation.get("preserveSourceSheets", True) is False:
        for source_sheet_name in source_sheet_names:
            if source_sheet_name == output_sheet_name:
                continue
            workbook.remove(workbook[source_sheet_name])

    return {
        "type": "merge_sheets",
        "sourceSheets": source_sheet_names,
        "outputSheetName": output_sheet_name,
    }


def build_split_sheet_name(source_sheet: str, output_prefix: Any, group_value: str) -> str:
    safe_group_name = sanitize_sheet_name(group_value or "Blank", "Blank")
    prefix = normalize_scalar(output_prefix) or f"{source_sheet} -"
    return sanitize_sheet_name(f"{prefix} {safe_group_name}", f"{source_sheet} Split")


def apply_split_sheet(workbook, operation: Dict[str, Any]):
    source_sheet_name = normalize_scalar(operation.get("sourceSheetName") or operation.get("sheetName"))
    if not source_sheet_name:
        raise ValueError("split_sheet operations require sourceSheetName")
    if source_sheet_name not in workbook.sheetnames:
        raise ValueError(f'split_sheet could not find source sheet "{source_sheet_name}"')

    by_column = normalize_scalar(operation.get("byColumn"))
    if not by_column:
        raise ValueError("split_sheet operations require byColumn")

    source_worksheet = workbook[source_sheet_name]
    header_row_index = detect_header_row_index(source_worksheet)
    headers = get_headers(source_worksheet, header_row_index)
    header_map = get_header_map(headers)
    split_column_index = header_map.get(normalize_column_name(by_column))
    if split_column_index is None:
        raise ValueError(f'Could not find split column "{by_column}"')

    grouped_row_indexes: Dict[str, List[int]] = {}
    for row_index in range(header_row_index + 1, source_worksheet.max_row + 1):
        group_value = normalize_scalar(source_worksheet.cell(row=row_index, column=split_column_index).value) or "Blank"
        grouped_row_indexes.setdefault(group_value, []).append(row_index)

    created_sheets = []
    for group_value, data_row_indexes in grouped_row_indexes.items():
        desired_name = build_split_sheet_name(
            source_sheet=source_sheet_name,
            output_prefix=operation.get("outputSheetPrefix"),
            group_value=group_value,
        )
        output_sheet_name = ensure_unique_sheet_name(list(workbook.sheetnames), desired_name)
        split_worksheet = workbook.create_sheet(title=output_sheet_name)
        copy_sheet_properties(source_worksheet, split_worksheet)
        prefix_and_header_rows = list(range(1, header_row_index + 1))
        source_row_indexes = prefix_and_header_rows + data_row_indexes
        copy_selected_rows_to_sheet(
            source_worksheet,
            split_worksheet,
            source_row_indexes,
            translate_formula_data_rows=set(data_row_indexes),
        )
        auto_filter_ref = build_autofilter_ref(len(headers), len(source_row_indexes))
        if auto_filter_ref:
            split_worksheet.auto_filter.ref = auto_filter_ref
        created_sheets.append(output_sheet_name)

    if operation.get("preserveSourceSheet", True) is False:
        workbook.remove(source_worksheet)

    return {
        "type": "split_sheet",
        "sourceSheetName": source_sheet_name,
        "byColumn": by_column,
        "createdSheets": created_sheets,
    }


def apply_operations(workbook, target_sheet_names: List[str], operations: List[Dict[str, Any]]):
    summaries = []
    for operation in operations:
        operation_type = normalize_scalar(operation.get("type"))
        if operation_type not in SUPPORTED_OPERATION_TYPES:
            raise UnsupportedOperationError(
                f'Python spreadsheet worker does not support operation "{operation_type}" yet'
            )

        if operation_type == "merge_sheets":
            summaries.append(apply_merge_sheets(workbook, operation))
            continue

        if operation_type == "split_sheet":
            summaries.append(apply_split_sheet(workbook, operation))
            continue

        explicit_sheet = normalize_scalar(operation.get("sheetName"))
        target_names = [explicit_sheet] if explicit_sheet else target_sheet_names
        if not target_names:
            raise ValueError(f'Operation "{operation_type}" did not resolve to any sheet')

        for sheet_name in target_names:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f'Operation "{operation_type}" could not find sheet "{sheet_name}"')
            worksheet = workbook[sheet_name]
            if operation_type == "add_column":
                apply_add_column(worksheet, operation)
            elif operation_type == "update_cells":
                apply_update_cells(worksheet, operation)
            elif operation_type == "add_row":
                apply_add_row(worksheet, operation)
            elif operation_type == "sort_rows":
                apply_sort_rows(worksheet, operation)
            elif operation_type == "add_totals_row":
                apply_add_totals_row(worksheet, operation)
            elif operation_type == "reorder_rows":
                apply_reorder_rows(worksheet, operation)

            summaries.append({"type": operation_type, "sheetName": sheet_name})

    return summaries


def serialize_workbook(workbook, output_format: str) -> Tuple[bytes, str]:
    if output_format == "csv":
        if len(workbook.sheetnames) != 1:
            raise ValueError("CSV output requires the transformed workbook to contain exactly one sheet")
        worksheet = workbook[workbook.sheetnames[0]]
        stream = io.StringIO()
        writer = csv.writer(stream)
        for row in worksheet.iter_rows(values_only=True):
            writer.writerow(["" if value is None else value for value in row])
        return stream.getvalue().encode("utf-8"), "text/csv"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return (
        buffer.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def validate_workbook_bytes(output_bytes: bytes, output_format: str):
    if output_format == "csv":
        try:
            output_bytes.decode("utf-8")
            return
        except UnicodeDecodeError as exc:
            raise ValueError("Generated CSV output is not valid UTF-8") from exc

    try:
        workbook = load_workbook(io.BytesIO(output_bytes), data_only=False, keep_vba=True)
    except Exception as exc:
        raise ValueError("Generated workbook could not be reopened") from exc

    if not workbook.sheetnames:
        raise ValueError("Generated workbook does not contain any sheets")


def build_final_sheet_summaries(workbook):
    summaries = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        header_row_index = detect_header_row_index(worksheet)
        headers = get_headers(worksheet, header_row_index)
        summaries.append(
            {
                "sheetName": sheet_name,
                "rowCount": max(worksheet.max_row - header_row_index, 0),
                "outputColumnCount": len(headers),
                "columns": headers,
                "previewRows": get_sheet_preview(worksheet, header_row_index, headers, 3),
            }
        )
    return summaries


def transform_workbook(payload: WorkerRequest):
    source_bytes = decode_workbook_buffer(payload)
    workbook = load_input_workbook(source_bytes, payload.sourceFilename)
    if not workbook.sheetnames:
        raise ValueError("Spreadsheet does not contain any sheets")

    keep_columns = normalize_string_list(payload.keepColumns)
    remove_columns = normalize_string_list(payload.removeColumns)
    redact_columns = normalize_string_list(payload.redactColumns)
    operations = normalize_operations(payload.operations)

    if not (keep_columns or remove_columns or redact_columns or operations):
        raise ValueError("At least one spreadsheet transformation must be requested")

    output_format = workbook_output_format(payload.sourceFilename, payload.outputFormat)
    target_sheet_names = resolve_target_sheets(workbook, normalize_string_list(payload.sheetNames))

    matched_columns = {"keep": set(), "remove": set(), "redact": set()}
    sheet_summaries = []
    warnings = []

    for sheet_name in target_sheet_names:
        worksheet = workbook[sheet_name]
        legacy_summary = apply_legacy_column_transforms(
            worksheet,
            keep_columns=keep_columns,
            remove_columns=remove_columns,
            redact_columns=redact_columns,
            redaction_text=payload.redactionText or DEFAULT_REDACTION_TEXT,
        )
        sheet_summaries.append(legacy_summary)
        matched_columns["keep"].update(legacy_summary["keptColumns"])
        matched_columns["remove"].update(legacy_summary["removedColumns"])
        matched_columns["redact"].update(legacy_summary["redactedColumns"])

    operation_summaries = apply_operations(workbook, target_sheet_names, operations)

    if payload.sourceFilename.lower().endswith(".xlsm"):
        warnings.append(
            "The Python spreadsheet worker loaded the workbook with VBA preservation enabled, but the transformed output is emitted as xlsx or csv."
        )

    output_bytes, mime_type = serialize_workbook(workbook, output_format)
    validate_workbook_bytes(output_bytes, output_format)

    return {
        "engine": "python_worker",
        "bufferBase64": base64.b64encode(output_bytes).decode("utf-8"),
        "bytes": len(output_bytes),
        "mimeType": mime_type,
        "filename": build_output_filename(payload.sourceFilename, output_format),
        "summary": {
            "engine": "python_worker",
            "outputFormat": output_format,
            "sheetCount": len(workbook.sheetnames),
            "sheets": sheet_summaries,
            "finalSheets": build_final_sheet_summaries(workbook),
            "operationsApplied": operation_summaries,
            "matchedColumns": {
                "keep": sorted(matched_columns["keep"]),
                "remove": sorted(matched_columns["remove"]),
                "redact": sorted(matched_columns["redact"]),
            },
            "warnings": warnings,
        },
    }


@app.get("/health", response_model=HealthResponse)
def health():
    return HealthResponse(ok=True, engine="python_worker")


@app.post("/inspect-workbook")
def inspect_workbook_endpoint(payload: WorkerRequest):
    try:
        workbook_bytes = decode_workbook_buffer(payload)
        return inspect_workbook(workbook_bytes, payload.sourceFilename, payload.maxPreviewRows)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail={
                "message": str(exc) or "Failed to inspect workbook",
                "code": "INSPECTION_FAILED",
            },
        ) from exc


@app.post("/apply-plan")
def apply_plan_endpoint(payload: WorkerRequest):
    try:
        return transform_workbook(payload)
    except UnsupportedOperationError as exc:
        raise HTTPException(
            status_code=422,
            detail={
                "message": str(exc),
                "code": "UNSUPPORTED_OPERATION",
            },
        ) from exc
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail={
                "message": str(exc) or "Failed to transform workbook",
                "code": "TRANSFORM_FAILED",
            },
        ) from exc
